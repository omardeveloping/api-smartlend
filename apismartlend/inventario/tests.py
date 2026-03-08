from datetime import datetime, timedelta

from io import BytesIO

from django.utils import timezone
from rest_framework import status
from rest_framework.test import APITestCase

from inventario.models import (
    categoria_herramienta,
    herramienta_individual,
    historial_herramienta,
    tipo_herramienta,
)
from operaciones.models import prestamo
from usuarios.models import Usuario, rol_usuarios


class HerramientaEstadoUsableTests(APITestCase):
    def setUp(self):
        self.categoria = categoria_herramienta.objects.create(nombre='Electricas')
        self.tipo = tipo_herramienta.objects.create(
            nombre='Taladro',
            descripcion='Taladro de prueba',
            id_categoria=self.categoria,
        )
        self.now = timezone.now()

    def _crear_herramienta(self, codigo, estado, disponible=True, numero_prestamos=0):
        return herramienta_individual.objects.create(
            marca='Bosch',
            modelo='X1',
            numero_prestamos=numero_prestamos,
            codigo_barras=codigo,
            estado_herramienta=estado,
            disponible=disponible,
            fecha_adquisicion=self.now,
            id_tipo_herramienta=self.tipo,
        )

    def test_herramienta_danada_o_defectuosa_se_mantiene_fuera_de_stock(self):
        self._crear_herramienta(
            'INV-BUENA',
            herramienta_individual.EstadoHerramienta.BUENO,
            disponible=True,
        )
        defectuosa = self._crear_herramienta(
            'INV-DEF',
            herramienta_individual.EstadoHerramienta.DEFECTUOSO,
            disponible=True,
        )

        defectuosa.refresh_from_db()
        self.tipo.refresh_from_db()

        self.assertFalse(defectuosa.disponible)
        self.assertEqual(self.tipo.stock, 1)

    def test_marcar_usable_cambia_estado_y_reactiva_disponibilidad(self):
        herramienta = self._crear_herramienta(
            'INV-DA',
            herramienta_individual.EstadoHerramienta.DANADO,
            disponible=True,
        )
        self.tipo.refresh_from_db()
        self.assertEqual(self.tipo.stock, 0)

        response = self.client.post(
            f'/inventario/api/herramientas/{herramienta.id_herramienta}/marcar-usable/',
            {'estado_herramienta': herramienta_individual.EstadoHerramienta.BUENO},
            format='json',
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        herramienta.refresh_from_db()
        self.tipo.refresh_from_db()
        self.assertEqual(
            herramienta.estado_herramienta,
            herramienta_individual.EstadoHerramienta.BUENO,
        )
        self.assertTrue(herramienta.disponible)
        self.assertEqual(self.tipo.stock, 1)

    def test_marcar_usable_rechaza_estado_no_usable(self):
        herramienta = self._crear_herramienta(
            'INV-DEF2',
            herramienta_individual.EstadoHerramienta.DEFECTUOSO,
            disponible=False,
        )

        response = self.client.post(
            f'/inventario/api/herramientas/{herramienta.id_herramienta}/marcar-usable/',
            {'estado_herramienta': herramienta_individual.EstadoHerramienta.DANADO},
            format='json',
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('detail', response.data)

    def test_no_usables_endpoint_lista_herramientas_no_usables(self):
        self._crear_herramienta(
            'INV-BUENA-2',
            herramienta_individual.EstadoHerramienta.BUENO,
            disponible=True,
        )
        self._crear_herramienta(
            'INV-DEF-3',
            herramienta_individual.EstadoHerramienta.DEFECTUOSO,
            disponible=True,
        )
        self._crear_herramienta(
            'INV-DAN-3',
            herramienta_individual.EstadoHerramienta.DANADO,
            disponible=True,
        )

        response = self.client.get('/inventario/api/herramientas/no-usables/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        codigos = {item['codigo_barras'] for item in response.data}
        self.assertEqual(codigos, {'INV-DEF-3', 'INV-DAN-3'})

    def test_no_usables_endpoint_permite_filtrar_por_estado(self):
        self._crear_herramienta(
            'INV-DEF-4',
            herramienta_individual.EstadoHerramienta.DEFECTUOSO,
            disponible=False,
        )
        self._crear_herramienta(
            'INV-DAN-4',
            herramienta_individual.EstadoHerramienta.DANADO,
            disponible=False,
        )

        response = self.client.get(
            '/inventario/api/herramientas/no-usables/?estado_herramienta=Defectuoso'
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['codigo_barras'], 'INV-DEF-4')

    def test_no_usables_endpoint_rechaza_estado_invalido(self):
        self._crear_herramienta(
            'INV-DEF-5',
            herramienta_individual.EstadoHerramienta.DEFECTUOSO,
            disponible=False,
        )

        response = self.client.get(
            '/inventario/api/herramientas/no-usables/?estado_herramienta=Bueno'
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('detail', response.data)

    def test_historial_endpoint_expone_prestador_y_receptor(self):
        rol_bodeguero = rol_usuarios.objects.create(
            nombre='bodeguero',
            desc='Bodeguero',
            permisos='inventario',
        )
        rol_alumno = rol_usuarios.objects.create(
            nombre='Alumno',
            desc='Alumno',
            permisos='prestamos',
        )
        bodeguero = Usuario.objects.create(
            correo='bodeguero-historial@example.com',
            rut='11111111-1',
            nombres='Bodega',
            apellidos='Uno',
            id_rol=rol_bodeguero,
        )
        receptor = Usuario.objects.create(
            correo='alumno-historial@example.com',
            rut='22222222-2',
            nombres='Alumno',
            apellidos='Dos',
            id_rol=rol_alumno,
        )
        herramienta = self._crear_herramienta(
            'INV-HIST-1',
            herramienta_individual.EstadoHerramienta.BUENO,
            disponible=False,
        )
        loan = prestamo.objects.create(
            fecha_prestamo=self.now - timedelta(hours=1),
            fecha_devolucion_esperada=self.now + timedelta(days=1),
            estado_prestamo=prestamo.EstadoPrestamo.ENTREGADO,
            id_usuario=receptor,
        )
        historial_herramienta.objects.create(
            herramienta=herramienta,
            estado_herramienta=herramienta.estado_herramienta,
            prestamo=loan,
            usuario=bodeguero,
        )

        response = self.client.get(
            f'/inventario/api/historial-herramientas/?herramienta={herramienta.id_herramienta}'
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        registro = response.data[0]
        self.assertEqual(registro['usuario_prestador_rut'], bodeguero.rut)
        self.assertEqual(registro['usuario_prestador_nombre'], 'Bodega Uno')
        self.assertEqual(registro['usuario_receptor_rut'], receptor.rut)
        self.assertEqual(registro['usuario_receptor_nombre'], 'Alumno Dos')

    def test_historial_exportar_excel_retorna_archivo_y_respeta_filtro(self):
        from openpyxl import load_workbook

        rol_bodeguero = rol_usuarios.objects.create(
            nombre='bodeguero',
            desc='Bodeguero',
            permisos='inventario',
        )
        rol_alumno = rol_usuarios.objects.create(
            nombre='Alumno',
            desc='Alumno',
            permisos='prestamos',
        )
        bodeguero = Usuario.objects.create(
            correo='bodeguero-excel@example.com',
            rut='33333333-3',
            nombres='Bodega',
            apellidos='Excel',
            id_rol=rol_bodeguero,
        )
        receptor = Usuario.objects.create(
            correo='alumno-excel@example.com',
            rut='44444444-4',
            nombres='Alumno',
            apellidos='Excel',
            id_rol=rol_alumno,
        )
        herramienta_1 = self._crear_herramienta(
            'INV-XLS-1',
            herramienta_individual.EstadoHerramienta.BUENO,
            disponible=False,
        )
        herramienta_2 = self._crear_herramienta(
            'INV-XLS-2',
            herramienta_individual.EstadoHerramienta.BUENO,
            disponible=False,
        )

        loan_1 = prestamo.objects.create(
            fecha_prestamo=self.now - timedelta(hours=2),
            fecha_devolucion_esperada=self.now + timedelta(days=1),
            estado_prestamo=prestamo.EstadoPrestamo.ENTREGADO,
            id_usuario=receptor,
        )
        loan_2 = prestamo.objects.create(
            fecha_prestamo=self.now - timedelta(hours=1),
            fecha_devolucion_esperada=self.now + timedelta(days=1),
            estado_prestamo=prestamo.EstadoPrestamo.ENTREGADO,
            id_usuario=receptor,
        )
        historial_herramienta.objects.create(
            herramienta=herramienta_1,
            estado_herramienta=herramienta_1.estado_herramienta,
            prestamo=loan_1,
            usuario=bodeguero,
        )
        historial_herramienta.objects.create(
            herramienta=herramienta_2,
            estado_herramienta=herramienta_2.estado_herramienta,
            prestamo=loan_2,
            usuario=bodeguero,
        )

        response = self.client.get(
            f'/inventario/api/historial-herramientas/exportar-excel/?herramienta={herramienta_1.id_herramienta}'
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn(
            'trazabilidad_herramientas.xlsx',
            response.get('Content-Disposition', ''),
        )

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(
            headers,
            [
                'id_historial',
                'codigo_barras',
                'tipo_herramienta',
                'estado_herramienta',
                'registrada_en',
                'prestamo_id',
                'prestamo_codigo',
                'usuario_prestador_nombre',
                'usuario_prestador_rut',
                'usuario_receptor_nombre',
                'usuario_receptor_rut',
            ],
        )
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet['B2'].value, 'INV-XLS-1')

    def test_top5_usadas_mes_excel_retorna_archivo_ordenado(self):
        from openpyxl import load_workbook

        rol = rol_usuarios.objects.create(
            nombre='Alumno',
            desc='Alumno',
            permisos='prestamos',
        )
        usuario = Usuario.objects.create(
            correo='usuario-top5@example.com',
            rut='55555555-5',
            nombres='Top',
            apellidos='Usuario',
            id_rol=rol,
        )

        inicio_mes = timezone.localdate().replace(day=1)
        fecha_mes_actual = timezone.make_aware(
            datetime.combine(inicio_mes + timedelta(days=10), datetime.min.time())
        )
        fecha_mes_anterior = fecha_mes_actual - timedelta(days=40)

        h1 = self._crear_herramienta(
            'INV-TOP-1',
            herramienta_individual.EstadoHerramienta.BUENO,
            disponible=True,
            numero_prestamos=10,
        )
        h2 = self._crear_herramienta(
            'INV-TOP-2',
            herramienta_individual.EstadoHerramienta.BUENO,
            disponible=True,
            numero_prestamos=7,
        )
        h3 = self._crear_herramienta(
            'INV-TOP-3',
            herramienta_individual.EstadoHerramienta.BUENO,
            disponible=True,
            numero_prestamos=3,
        )

        def crear_prestamo(fecha, herramienta):
            loan = prestamo.objects.create(
                fecha_prestamo=fecha,
                fecha_devolucion_esperada=fecha + timedelta(days=1),
                estado_prestamo=prestamo.EstadoPrestamo.FINALIZADO,
                id_usuario=usuario,
            )
            loan.herramientas.add(herramienta)

        for _ in range(3):
            crear_prestamo(fecha_mes_actual, h1)
        for _ in range(2):
            crear_prestamo(fecha_mes_actual + timedelta(hours=1), h2)
        crear_prestamo(fecha_mes_actual + timedelta(hours=2), h3)
        for _ in range(4):
            crear_prestamo(fecha_mes_anterior, h2)

        response = self.client.get('/inventario/api/herramientas/top5-usadas-mes-excel/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('top5_herramientas_', response.get('Content-Disposition', ''))

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(
            headers,
            [
                'posicion',
                'id_herramienta',
                'codigo_barras',
                'tipo_herramienta',
                'marca',
                'modelo',
                'usos_mes',
                'numero_prestamos',
            ],
        )

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[0][2], 'INV-TOP-1')
        self.assertEqual(rows[0][6], 3)
        self.assertEqual(rows[1][2], 'INV-TOP-2')
        self.assertEqual(rows[1][6], 2)
        self.assertEqual(rows[2][2], 'INV-TOP-3')
        self.assertEqual(rows[2][6], 1)
